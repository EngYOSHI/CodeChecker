import os
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.numbers import FORMAT_TEXT
import re

import common as c


_ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


def write_xl(students: list[c.Student]):
    wb = openpyxl.Workbook()
    wb.properties.creator = 'CodeChecker'
    wb.properties.lastModifiedBy = 'CodeChecker'
    ws = wb["Sheet"]
    ws.append(["学籍番号", "課題番号", "ｺﾝﾊﾟｲﾙ\n結果", "コンパイル備考", "コンパイルログ",
               "ﾃｽﾄｹｰｽ", "テスト\n結果", "テスト結果備考", "ﾃｽﾄｹｰｽ\n一致率", "出力種類", "出力"])
    ws.row_dimensions[1].height = 27
    # 先頭行にスタイル適用
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal = "center",
                                   vertical = "center", wrapText = True)
        cell.font = Font(bold = True)
    # 列幅を設定
    column_width = {"A":11, "B":10, "C":10, "D":20, "E":20,
                    "F":10, "G":10, "H":23, "I":10, "J":10, "K":30}
    for column, width in column_width.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2" #先頭行を固定
    ws.auto_filter.ref = "A1:K1"  # フィルタを設定
    wraptext = ["D", "E", "H", "K"]
    row = 2
    for student in students:
        for task_results in student.task_results:
            task = task_results.task
            compile_result = task_results.compile_result
            if task_results.run_results is None:
                # コンパイルエラーの場合
                write_common(ws, row, student.student_number, task.tasknumber,
                             compile_result.result, compile_result.reason,
                             compile_result.stdout)
                row += 1
                continue
            for testcase_number, run_result in enumerate(task_results.run_results):
                write_common(ws, row, student.student_number, task.tasknumber,
                            compile_result.result, compile_result.reason,
                            compile_result.stdout)
                ws["F" + str(row)].value = task.tasknumber + f" [{str(testcase_number + 1)}]"
                ws["G" + str(row)].value = run_result.result.value
                cellfill(ws["G" + str(row)],
                         [(c.RunResultState.OK, "00b050"),
                          (c.RunResultState.NG, "e09694"),
                          (c.RunResultState.SKIP, "c5d9f1"),
                          (c.RunResultState.ENCERR, "a2a2a2"),
                          (c.RunResultState.NOTEST, "a2a2a2")]
                        )
                ws["G" + str(row)].alignment = Alignment(horizontal = "center")
                run_result_reason = "" if run_result.reason is None else run_result.reason.value
                ws["H" + str(row)].value = valconv(run_result_reason, str, none="")
                if run_result.ratio is None and run_result.reason == c.RunResultReason.WRONG:
                    # 一致率計算がタイムアウトした場合
                    ws["I" + str(row)].value = "ﾀｲﾑｱｳﾄ"
                else:
                    ws["I" + str(row)].value = valconv(run_result.ratio, float, none="")
                    ws["I" + str(row)].number_format = "0.000"
                if task.testcases is not None:
                    ws["J" + str(row)].value = task.testcases[testcase_number].out_type.value
                ws["K" + str(row)].value = str_escape(run_result.str_out)
                row += 1
    # 全体的な設定
    for row2 in range(2, row):
        # 折り返して全体を表示
        for column in wraptext:
            ws[column + str(row2)].alignment = Alignment(
                wrapText = True, vertical = 'top')
        ws.row_dimensions[row2].height = 13.5  # 高さ調節
    path = os.path.join(c.RESULT_PATH, "result.xlsx")
    wb.save(path)


def write_common(ws, row: int, student_number: str, task_number: str,
                 compile_result: bool, compile_reason: str | None, compile_stdout: str):
    ws["A" + str(row)].value = student_number
    ws["A" + str(row)].number_format = FORMAT_TEXT
    ws["B" + str(row)].value = task_number
    ws["B" + str(row)].number_format = FORMAT_TEXT
    ws["C" + str(row)].value = valconv(compile_result, bool, "OK", "NG")
    cellfill(ws["C" + str(row)], [("OK", "00b050"),("NG", "e09694")])
    ws["C" + str(row)].alignment = Alignment(horizontal = "center")
    ws["D" + str(row)].value = valconv(compile_reason, str, none="")
    ws["E" + str(row)].value = str_escape(compile_stdout)


def str_escape(s: str | None) -> str:
    """不正な文字が含まれているとIllegalCharacterError例外を出して止まるので?に置き換え"""
    if s is None:
        return ""
    else:
        return _ILLEGAL_CHARACTERS_RE.sub("?", s)


def valconv(data, type, str_true = "True", str_false = "False", none = "None"):
    if data is None: return none
    elif type == bool:
        if data: return str_true
        else: return str_false
    else: return data


def cellfill(cell, conditions):
    #conditionsは(文字列, 色)のリスト
    for s, color in conditions:
        if cell.value == s:
            cell.fill = PatternFill(patternType='solid', fgColor=color)
            return
