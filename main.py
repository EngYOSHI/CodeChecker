import subprocess
import os
import sys
import re
import time
import pprint
import difflib
import argparse
import openpyxl
import shutil

DEBUG = False
PRINT_SCORE = True
GCC_PATH = "mingw64\\bin\\"
SRC_PATH = "src\\"
WORK_PATH = "work\\"
CASE_PATH = "case\\"
RESULT_PATH = "result\\"
TEMP_PATH = "temp\\"
TIMEOUT = 5
NOCOLOR = False


srclist = []


def main():
    students = file2list(os.path.join(CASE_PATH, "students.txt"))
    print("学生数: " + str(len(students)))
    debug(students)
    tasks = read_taskfiles(os.path.join(CASE_PATH, "tasks.txt"))
    print("課題数: " + str(len(tasks)))
    tasks = read_casefiles(tasks)
    debug(pprint.pformat(tasks, sort_dicts=False))
    src_listgen()
    print("ソースファイル数: " + str(len(srclist)))
    debug(srclist)
    res = eval_loop(students, tasks)
    del_temp()  # tempフォルダを削除
    #pprint.pprint(res, sort_dicts=False)
    print("未処理のファイル数: " + str(len(srclist)))
    print("結果をxlsxファイルに書き込み中...")
    write_xl(res)
    write_untouched()


def write_untouched():
    with open(os.path.join(RESULT_PATH, "untouched.txt"), 'w') as f:
        for x in srclist:
            f.write(x + "\n")


def write_xl(res):
    #[{student:str,
    #  result:[{task:str,
    #           compile:{result:bool, reason:None|str, stdout:str|None},
    #           run:[{result:bool|None, reason:None|str, output:str|None, ratio:float|None} | None]
    #         }]
    #}]
    global RESULT_PATH
    wb = openpyxl.Workbook()
    wb.properties.creator = 'CodeChecker'
    wb.properties.lastModifiedBy = 'CodeChecker'
    ws = wb["Sheet"]
    ws.append(["学籍番号", "課題番号", "ｺﾝﾊﾟｲﾙ結果", "コンパイル備考", "コンパイルログ", "ﾃｽﾄｹｰｽ", "テスト結果", "テスト結果備考", "ﾃｽﾄｹｰｽ一致率", "標準出力"])
    for char in list("ABDEHJ"):
        ws.column_dimensions[char].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["F"].width = 7
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["I"].width = 13
    ws.freeze_panes = "A2" #先頭行を固定
    row = 2
    for res_student in res:
        ws.cell(row=row, column=1).value = res_student["student"]
        for task in res_student["result"]:
            ws.cell(row=row, column=2).value = task["task"]
            ws.cell(row=row, column=3).value = valconv(task["compile"]["result"], bool, t = "OK", f = "NG")
            cellfill(ws.cell(row=row, column=3), [("OK", "00b050"),("NG", "e09694")])
            ws.cell(row=row, column=4).value = valconv(task["compile"]["reason"], str, none="")
            ws.cell(row=row, column=5).value = valconv(task["compile"]["stdout"], str, none="")
            if task["run"] is None:
                ws.cell(row=row, column=6).value = "　" #コンパイルログが右のセルにはみ出さないように
                row += 1
                continue
            for i, run in enumerate(task["run"]):
                ws.cell(row=row, column=6).value = task["task"] + "_" + str(i + 1)
                ws.cell(row=row, column=7).value = valconv(run["result"], bool, t="OK", f="NG", none="SKIP")
                cellfill(ws.cell(row=row, column=7), [("OK", "00b050"),("NG", "e09694"),("SKIP", "c5d9f1")])
                ws.cell(row=row, column=8).value = valconv(run["reason"], str, none="")
                ws.cell(row=row, column=9).value = valconv(run["ratio"], float, none="")
                ws.cell(row=row, column=10).value = valconv(run["output"], str, none="")
                row += 1
    path = os.path.join(RESULT_PATH, "result.xlsx")
    while True:
        try:
            wb.save(path)
        except PermissionError:
            print(str_color(Color.RED, f'File "{path}" is open! Please close the file.'))
            print("Retry in 5 seconds...")
            time.sleep(5)
            continue
        break


def valconv(data, type, t = "True", f = "False", none = "None"):
    if data is None: return none
    elif type == bool:
        if data: return t
        else: return f
    else: return data


def cellfill(cell, conditions):
    #conditionsは(文字列, 色)のリスト
    for s, color in conditions:
        if cell.value == s:
            cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=color)
            return


def eval_loop(students, tasks):
    res = []
    for i, s in enumerate(students):
        res_student = []
        for t in tasks:
            r = eval(s, t)
            res_student += [r]
        res += [{"student":s, "result":res_student}]
        if PRINT_SCORE:
            progress = f"({i + 1}/{len(students)})"
            print_score(s, res_student, progress)
    return res


def eval(student, task):
    r = {"task":task["name"], "compile":None, "run":None}
    exe = f"{student}_{task["name"]}"
    debug("\n")
    debug(exe, "eval")
    debug("----------------------")
    src = get_latest_src(student, task["name"])
    debug(src, "srcfile")
    # コンパイル
    temp_reset()  # Tempフォルダの初期化
    r["compile"] = compile(src, exe)
    debug(r["compile"], "compile")
    if r["compile"]["result"]:
        # コンパイル成功ならテスト実行
        r["run"] = run_loop(exe + ".exe", task)
        debug("", "run_loop")
        debug(pprint.pformat(r["run"], sort_dicts=False))
        mv_temp2bin(exe + ".exe")  # 実行が終わったバイナリはbinフォルダへ移動
    return r


def get_latest_src(student, taskname):
    #命名条件に合致するファイルが一つもなければNoneを返す
    #命名条件に合致するファイルがあれば，評価対象のファイル名を返す
    global srclist
    r = student + "_" + taskname + r"(\([1-9][0-9]*\))?" + ".c"
    src = [x for x in srclist if re.match(r, x)]
    #命名条件に合致するファイルが一つもない -> None
    if len(src) == 0:
        return None
    #チェックしたファイルをソースファイルリストから削除する
    for s in src:
        srclist.remove(s)
    #チェックすべきソースファイルを識別
    #最新の提出物は，拡張子の前に番号が「付かない」，ピュアなファイル名のもの
    return f"{student}_{taskname}.c"


def compile(src, exe):
    res = {"result":True, "reason":None, "stdout":None}
    if src is None:
        res["result"] = False
        res["reason"] = "未提出orﾌｧｲﾙ名間違い"
        return res
    src_abs = os.path.join(os.path.abspath(SRC_PATH), src)
    exe_abs = os.path.join(os.path.abspath(TEMP_PATH), exe)
    cmd = ["gcc.exe", src_abs, "-o", exe_abs]
    r = subprocess.run(cmd, cwd=GCC_PATH, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,shell=True)
    res["stdout"] = byte2str(r.stdout)
    if r.returncode != 0:
        #コンパイル失敗
        res["result"] = False
        res["reason"] = "コンパイルエラー"
        if res["stdout"] is None:
            res["reason"] += " + 未サポートｴﾝｺｰﾄﾞ"
    return res


def run_loop(exe, task):
    res = []
    taskfn = task["name"] + "_" + str(task["count"]) + "_"
    for case in task["case"]:
        res += [run_exe(exe, taskfn, case)]
    return res


def run_exe(exe, taskfn, case):
    cmd = exe
    if case["arg"] is not None:
        debug(case["arg"], "arg")
        cmd = cmd + " " + case["arg"]
    res = {"result":None, "reason":None, "output":None, "ratio":None}
    proc = subprocess.Popen(cmd, cwd=TEMP_PATH, stdout=subprocess.PIPE, stdin=subprocess.PIPE, shell=True)
    try:
        if case["in"] is None:
            r = proc.communicate(timeout=TIMEOUT)
        else:
            r = proc.communicate(timeout=TIMEOUT, input=case["in"].encode("cp932"))
    except subprocess.TimeoutExpired:
        #タイムアウト
        proc.kill()
        debug("Time Out.", "run")
        res["result"] = False
        res["reason"] = "タイムアウト"
        return res
    res["output"] = byte2str(r[0])  #標準出力のバイトストリームを文字列に変換
    if case["out"] is not None:
        if res["output"] is None:
            res["result"] = False
            res["reason"] = "未サポートｴﾝｺｰﾄﾞ"
            return res
        res["ratio"] = round(difflib.SequenceMatcher(None, res["output"], case["out"], False).ratio(), 3)
        if res["output"] == case["out"]:
            res["result"] = True
        else:
            res["reason"] = "ﾃｽﾄｹｰｽと不一致"
            res["result"] = False
    return res


def print_score(s, res_student, progress):
    output = "Student No.: " + s + "   " + progress
    for r in res_student:
        output += "\n\tTask: " + r["task"]
        output += "\n\t\tCompile: " + bool2str(r["compile"]["result"], "OK", "NG")
        if r["compile"]["result"]:
            correct = 0
            failed = 0
            skip = 0
            for i in range(len(r["run"])):
                output += "\n\t\t[" + str(i) + "] -> Result: " + bool2str(r["run"][i]["result"], "OK", "NG", "SKIP")
                if r["run"][i]["result"] is None:
                    skip += 1
                elif r["run"][i]["result"] == True:
                    correct += 1
                else:
                    failed += 1
                    if r["run"][i]["reason"] is not None:
                        output += f" (Reason: {r['run'][i]['reason']})"
                    if r["run"][i]["ratio"] is not None:
                        output += f" (Ratio: {r['run'][i]['ratio']})"
            output += f"\n\t\tSummary: {correct}/{correct + failed}  (skip:{skip})"
        else:
            output += " (" + r["compile"]["reason"] + ")"
    print(output + "\n")


def mv_temp2bin(exe):
    """exeをtempからresult\binに移動
    
    既に存在する場合は上書き
    """
    bin_path = os.path.join(RESULT_PATH, "bin\\")
    if not os.path.isdir(bin_path):
        os.mkdir(bin_path)
    mv_from = os.path.join(TEMP_PATH, exe)
    mv_to = os.path.join(bin_path, exe)
    debug(f"{mv_from} -> {mv_to}", "mv_temp2bin")
    shutil.move(mv_from, mv_to)


def del_temp():
    """一時保存フォルダ(temp)がある場合削除する"""
    if os.path.isdir(TEMP_PATH):
        shutil.rmtree(TEMP_PATH)


def temp_reset():
    """一時保存フォルダ(temp)を初期化する
    
    一時保存フォルダが存在する場合はフォルダごと消す
    その後，workフォルダをtempフォルダとしてコピー
    """
    del_temp()
    debug(f"Copying working files: {WORK_PATH} -> {TEMP_PATH}", "temp_reset")
    shutil.copytree(WORK_PATH, TEMP_PATH)


def chkpath():
    if not os.path.isdir(GCC_PATH):
        error(f"コンパイラがないか，正しく配置されていません．({GCC_PATH})")
    if not os.path.isdir(SRC_PATH):
        error(f"ソースファイルの格納先({SRC_PATH})がありません．")
    if not os.path.isdir(WORK_PATH):
        error(f"読み込ませるファイルの格納先({WORK_PATH})がありません．")
    if not os.path.isfile(os.path.join(CASE_PATH, "students.txt")):
        error(f"学籍番号ファイル'students.txt'が{CASE_PATH}にありません．")
    if not os.path.isfile(os.path.join(CASE_PATH, "tasks.txt")):
        error(f"タスクファイル'tasks.txt'が{CASE_PATH}にありません．")
    if not os.path.isdir(RESULT_PATH):
        error(f"結果ファイルの格納先({RESULT_PATH})がありません．")


def file2list(filename):
    f = open(filename, 'r', encoding="utf-8")
    l = f.readlines()
    f.close()
    for i in range(len(l)):
        if l[i].strip() == "":
            l.pop(i)
        else:
            l[i] = l[i].rstrip("\n")
    return l


def file2str(filename):
    f = open(filename, 'r', encoding="utf-8")
    s = f.read()
    f.close()
    return s


def read_taskfiles(filename):
    l = file2list(filename)
    for i in range(len(l)):
        if re.fullmatch(r'[1-9][0-9]*-((A[1-9])|([1-9][a-z]*)) [1-9]', l[i]):
            temp = l[i].split(" ")
            l[i] = {"name" : temp[0], "count" : int(temp[1]), "case" : []}
        else:
            error("'tasks.txt'の構文エラー")
    return l


def read_casefiles(l):
    for i in range(len(l)):
        count = l[i]["count"]
        for j in range(count):
            d = {"arg":None, "out":None, "in":None}
            file_arg = os.path.join(CASE_PATH, l[i]["name"] + "_" + str(j) + "_arg.txt")
            file_out = os.path.join(CASE_PATH, l[i]["name"] + "_" + str(j) + "_out.txt")
            file_in = os.path.join(CASE_PATH, l[i]["name"] + "_" + str(j) + "_in.txt")
            if os.path.isfile(file_arg):
                d["arg"] = file2list(file_arg)[0]
            if os.path.isfile(file_out):
                d["out"] = file2str(file_out)
            if os.path.isfile(file_in):
                d["in"] = file2str(file_in)
            l[i]["case"] += [d]
    return l


def byte2str(byte) -> str:
    try:
        #UTF8でデコードする
        s = byte.decode("utf-8")
        s = s.replace("\r\n","\n")
        debug("utf-8", "byte2str")
        return s
    except:
        pass
    try:
        #SJISでデコードする
        s = byte.decode("cp932")
        s = s.replace("\r\n","\n")
        debug("cp932", "byte2str")
        return s
    except:
        pass
    debug("サポートされていないエンコード", "byte2str")
    return None


def bool2str(b, str_true, str_false, none = None):
    if b is None:
        return str_color(Color.BG_CYAN, none)
    elif b:
        return str_color(Color.BG_BLUE, str_true)
    else:
        return str_color(Color.BG_RED, str_false)


def error(msg, need_exit: bool = True):
    print(str_color(Color.RED, str(msg)))
    if need_exit:
        sys.exit(1)


def debug(msg, title = None):
    if DEBUG:
        if title is not None:
            print(str_color(Color.BG_GREEN, title + ":") + str_color(Color.GREEN, " " + str(msg)))
        else:
            print(str_color(Color.GREEN, str(msg)))


def chkarg():
    global DEBUG, SRC_PATH, WORK_PATH, CASE_PATH, RESULT_PATH, TEMP_PATH, TIMEOUT, NOCOLOR
    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--debug", help="デバッグ出力を有効にする", action='store_true')
    parser.add_argument('--src', help="ソースファイルの格納先を指定", type=str, default=SRC_PATH)
    parser.add_argument('--work', help="読み込みファイルの格納先を指定", type=str, default=WORK_PATH)
    parser.add_argument('--case', help="テストケースの格納先を指定", type=str, default=CASE_PATH)
    parser.add_argument('--result', help="結果の格納先を指定", type=str, default=RESULT_PATH)
    parser.add_argument('--temp', help="一時ファイル格納先を指定", type=str, default=TEMP_PATH)
    parser.add_argument('--timeout', help="1プログラム当たりのタイムアウト時間を秒で指定", type=int, default=TIMEOUT)
    parser.add_argument('--nocolor', help="色付き出力を無効化", action='store_true')
    args = parser.parse_args()
    DEBUG = args.debug
    SRC_PATH = args.src
    WORK_PATH = args.work
    CASE_PATH = args.case
    RESULT_PATH = args.result
    TEMP_PATH = args.temp
    TIMEOUT = args.timeout
    NOCOLOR = args.nocolor
    debug(f"Source: {SRC_PATH}", "chkarg")
    debug(f"Work: {WORK_PATH}", "chkarg")
    debug(f"Case: {CASE_PATH}", "chkarg")
    debug(f"Result: {RESULT_PATH}", "chkarg")
    debug(f"Temp: {TEMP_PATH}", "chkarg")
    debug(f"タイムアウト: {TIMEOUT}s", "chkarg")


def src_listgen():
    global SRC_PATH, srclist
    srclist = [
        f for f in os.listdir(SRC_PATH) if os.path.join(os.path.join(SRC_PATH, f))
    ]
    return srclist


class Color:
    RESET = "\033[0m"
    RED = "\033[31m"
    GREEN = "\033[32m"
    YELLOW = "\033[33m"
    BG_RED = "\033[41m"
    BG_GREEN = "\033[42m"
    BG_BLUE = "\033[44m"
    BG_CYAN = "\033[46m"


def str_color(color: Color, s: str):
    global NOCOLOR
    if NOCOLOR:
        return s
    else:
        return color + s + Color.RESET


if __name__ == "__main__":
    os.system("")
    chkarg()
    chkpath()
    main()